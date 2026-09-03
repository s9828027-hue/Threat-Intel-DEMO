"""
ThreatGate - Diff-to-Excel Report Module
差異摘要轉 Excel 審核報表

Input:  data/diff_summary.json
Output: data/diff_report.xlsx  - a reviewer-friendly, filterable report
"""

import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.environ.get("THREATGATE_DATA_DIR", "data")
DIFF_FILE = os.path.join(DATA_DIR, "diff_summary.json")
OUTPUT_FILE = os.path.join(DATA_DIR, "diff_report.xlsx")

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E5A")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
LARGE_NET_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_entries_sheet(ws, entries, include_large_network_flag=False):
    headers = ["Indicator", "Type", "Sources", "Categories", "First Seen"]
    if include_large_network_flag:
        headers.append("Large network?")

    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for r, entry in enumerate(entries, start=2):
        ws.cell(row=r, column=1, value=entry.get("indicator", ""))
        ws.cell(row=r, column=2, value=entry.get("type", ""))
        ws.cell(row=r, column=3, value=", ".join(entry.get("sources", [])))
        ws.cell(row=r, column=4, value=", ".join(entry.get("categories", []) or []))
        ws.cell(row=r, column=5, value=entry.get("first_seen") or "")

        is_large = entry.get("large_network", False)
        if include_large_network_flag:
            ws.cell(row=r, column=6, value="Yes" if is_large else "")

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER
            if is_large:
                cell.fill = LARGE_NET_FILL

    if entries:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(entries) + 1}"

    widths = [22, 10, 24, 20, 22]
    if include_large_network_flag:
        widths.append(14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def run(diff_file: str = None, output_file: str = None) -> str:
    diff_file = diff_file or DIFF_FILE
    output_file = output_file or OUTPUT_FILE

    with open(diff_file, "r", encoding="utf-8") as f:
        diff = json.load(f)

    wb = openpyxl.Workbook()

    ov = wb.active
    ov.title = "Summary"
    ov["B2"] = "Threat Intel Review Report"
    ov["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color="1F4E5A")
    ov["B3"] = f"Generated at: {diff.get('generated_at', '')}"
    ov["B3"].font = Font(name=FONT_NAME, size=10, color="666666")

    rows = [
        ("First run?", "Yes" if diff.get("is_first_run") else "No"),
        ("Baseline count", diff.get("baseline_count", 0)),
        ("Today's count", diff.get("today_count", 0)),
        ("Added", diff.get("added_count", 0)),
        ("Removed", diff.get("removed_count", 0)),
        ("Unchanged", diff.get("unchanged_count", 0)),
        ("Anomaly flag", "Yes" if diff.get("is_anomaly") else "No"),
        ("Anomaly reason", diff.get("anomaly_reason", "")),
    ]
    r = 5
    for label, val in rows:
        lc = ov.cell(row=r, column=2, value=label)
        lc.font = Font(name=FONT_NAME, size=10, bold=True)
        vc = ov.cell(row=r, column=3, value=val)
        vc.font = Font(name=FONT_NAME, size=10,
                        color="C00000" if label == "Anomaly flag" and diff.get("is_anomaly") else "000000")
        r += 1

    ov.cell(row=r + 1, column=2,
            value="Note: yellow rows are 'large network' flags for manual review, not auto-exclusions.").font = \
        Font(name=FONT_NAME, size=9, italic=True, color="666666")

    ov.column_dimensions["B"].width = 16
    ov.column_dimensions["C"].width = 50
    ov.sheet_view.showGridLines = False

    ws_added = wb.create_sheet("Added")
    write_entries_sheet(ws_added, diff.get("added_entries", []), include_large_network_flag=True)

    ws_removed = wb.create_sheet("Removed")
    write_entries_sheet(ws_removed, diff.get("removed_entries", []), include_large_network_flag=False)

    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)
    wb.save(output_file)
    print(f"Report written to {output_file}")
    return output_file


if __name__ == "__main__":
    run()
